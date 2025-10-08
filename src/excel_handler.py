"""
Manejo del archivo Excel para añadir resultados.
Este módulo se encarga de toda la interacción con el archivo Excel,
incluyendo creación, lectura, escritura y actualización de datos.
"""

# Importación de módulos necesarios
import openpyxl  # Para manipular archivos Excel
from openpyxl import Workbook  # Para crear nuevos archivos Excel
import os  # Para operaciones con el sistema de archivos
from config import EXCEL_FILE  # Importa la ruta del archivo Excel desde la configuración


class ExcelHandler:
    def __init__(self, filepath=EXCEL_FILE):
        """
        Constructor de la clase ExcelHandler.
        Args:
            filepath (str): Ruta del archivo Excel a manejar. Por defecto usa EXCEL_FILE de config.
        """
        self.filepath = filepath  # Almacena la ruta del archivo Excel
        self.workbook = None     # Referencia al libro de Excel (se inicializa en None)
        self.sheet = None        # Referencia a la hoja activa (se inicializa en None)
        # Headers base que siempre estarán presentes en el Excel
        self.base_headers = ['Date', 'UVigoProfesor', 'USCEmprego']
        
    def load_or_create(self):
        """
        Carga un archivo Excel existente o crea uno nuevo si no existe.
        """
        if os.path.exists(self.filepath):  # Verifica si el archivo existe
            # Si existe, lo carga
            self.workbook = openpyxl.load_workbook(self.filepath)
            self.sheet = self.workbook.active  # Obtiene la hoja activa
            print(f"Excel cargado: {self.filepath}")
        else:
            # Si no existe, crea uno nuevo
            self.workbook = Workbook()
            self.sheet = self.workbook.active
            self.sheet.title = "Resultados Scraper"  # Establece el nombre de la hoja
            print(f"Nuevo Excel creado: {self.filepath}")
    
    def _get_current_headers(self):
        """
        Obtiene los headers actuales del Excel desde la primera fila.
        
        Returns:
            list: Lista de headers actuales o lista vacía si no hay.
        """
        headers = []
        if self.sheet.max_row >= 1:  # Si hay al menos una fila
            for col in range(1, self.sheet.max_column + 1):
                header = self.sheet.cell(1, col).value
                if header:  # Solo añade headers no vacíos
                    headers.append(header)
        return headers
    
    def setup_headers(self, additional_names=None):
        """
        Configura los headers del Excel con la estructura requerida.
        Estructura: Date | UVigoProfesor | USCEmprego | [otros nombres de config]
        
        Args:
            additional_names (list): Nombres adicionales de sitios web desde urls_config.json
        """
        current_headers = self._get_current_headers()
        
        # Si no hay headers, crear la estructura completa
        if not current_headers:
            # Comenzar con los headers base
            all_headers = self.base_headers.copy()
            
            # Añadir nombres adicionales de sitios web (evitando duplicados)
            if additional_names:
                for name in additional_names:
                    if name not in all_headers:
                        all_headers.append(name)
            
            # Escribir todos los headers en la primera fila
            for col, header in enumerate(all_headers, start=1):
                self.sheet.cell(1, col, header)
            
            print(f"Headers configurados: {all_headers}")
            
        else:
            # Si ya hay headers, verificar si necesitamos añadir nuevos
            if additional_names:
                next_col = len(current_headers) + 1
                for name in additional_names:
                    if name not in current_headers:
                        self.sheet.cell(1, next_col, name)
                        current_headers.append(name)
                        next_col += 1
                        print(f"Nuevo header añadido: {name}")
    
    def update_single_row(self, results):
        """
        Actualiza la única fila de datos (fila 2) con los nuevos resultados.
        Reemplaza completamente los datos anteriores.
        
        Args:
            results (dict): Diccionario con los resultados del scraping.
                          Debe incluir 'date' y los valores para cada sitio web.
        """
        if not results:
            print("No hay resultados para actualizar")
            return
        
        # Obtener headers actuales
        current_headers = self._get_current_headers()
        
        if not current_headers:
            print("Error: No hay headers configurados")
            return
        
        # Limpiar la fila 2 completamente
        for col in range(1, len(current_headers) + 1):
            self.sheet.cell(2, col, "")
        
        # Llenar la fila 2 con los nuevos datos
        for col, header in enumerate(current_headers, start=1):
            if header == 'Date':
                # Para la columna Date, usar el timestamp del resultado
                value = results.get('date', '')
            else:
                # Para las demás columnas, usar el valor correspondiente al nombre del sitio
                value = results.get(header, 'NO')  # Default 'NO' si no se encuentra
            
            self.sheet.cell(2, col, value)
        
        print(f"Fila de datos actualizada con fecha: {results.get('date', 'N/A')}")
    
    def append_results(self, results):
        """
        Procesa los resultados del scraping y actualiza el Excel.
        En lugar de añadir filas, actualiza la estructura y la única fila de datos.
        
        Args:
            results (list): Lista de diccionarios con los datos del scraping.
                          Se espera solo un diccionario con todos los resultados.
        """
        if not results:  # Verifica si hay resultados para procesar
            print("No hay resultados para añadir")
            return
        
        # Tomar el primer (y único) resultado
        result = results[0] if isinstance(results, list) else results
        
        # Extraer nombres de sitios web del resultado (excluyendo 'date')
        site_names = [key for key in result.keys() if key != 'date']
        
        # Configurar headers (incluyendo cualquier sitio nuevo)
        self.setup_headers(additional_names=site_names)
        
        # Actualizar la única fila de datos
        self.update_single_row(result)
        
        print("Excel actualizado con nueva estructura y datos")
    
    def save(self):
        """
        Guarda los cambios en el archivo Excel.
        Crea el directorio si no existe.
        """
        # Asegura que el directorio exista antes de guardar
        # exist_ok=True evita errores si el directorio ya existe
        os.makedirs(os.path.dirname(self.filepath), exist_ok=True)
        
        # Guarda el workbook en el archivo especificado
        self.workbook.save(self.filepath)
        print(f"Excel guardado: {self.filepath}")
    
    def close(self):
        """
        Cierra el archivo Excel para liberar recursos.
        Importante llamar a este método al finalizar para evitar
        corrupción de archivos y liberar memoria.
        """
        # Verifica que el workbook esté abierto antes de cerrarlo
        if self.workbook:
            self.workbook.close()  # Cierra el workbook y libera recursos


def update_excel_with_results(results):
    """
    Función auxiliar que encapsula todo el proceso de actualización del Excel.
    Esta es la función principal que se debe usar desde otros módulos.
    
    Args:
        results (list): Lista con un diccionario que contiene los resultados del scraping.
                       El diccionario debe incluir 'date' y los resultados de cada sitio web.
    """
    # Crea una instancia del manejador de Excel
    handler = ExcelHandler()
    
    try:
        # Carga el archivo existente o crea uno nuevo
        handler.load_or_create()
        
        # Procesa y actualiza los resultados (estructura + datos)
        handler.append_results(results)
        
        # Guarda los cambios en el archivo
        handler.save()
        
    finally:
        # Asegura que el archivo se cierre correctamente
        # El bloque finally se ejecuta siempre, incluso si hay errores
        handler.close()


if __name__ == "__main__":
    # Código de prueba que se ejecuta solo cuando este archivo
    # se ejecuta directamente (no cuando se importa como módulo)
    
    # Crea datos de prueba en el formato esperado por el nuevo sistema
    test_results = [ 
        {
            'date': '2025-10-08',  # Fecha de ejecución
            'UVigoProfesor': 'YES',  # Resultado del scraping de UVigo
            'USCEmprego': 'NO'  # Resultado del scraping de USC
        }
    ]
    
    # Ejecuta la función de actualización con los datos de prueba
    update_excel_with_results(test_results)